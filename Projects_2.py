from werkzeug.utils import secure_filename
from flask import Flask,app,request,g,render_template,redirect,url_for,abort,send_file
from flask import jsonify,make_response,send_from_directory,flash
from contextlib import closing
import zipfile
import sqlite3
import os
import base64
import pandas as pd
from dbfread import DBF
import openpyxl
import numpy as np
import shutil

app = Flask(__name__)

UPLOAD_FOLDER = 'C:\\Users\\minghsuantu\\Desktop\\flask try\\static\\upload'
app.secret_key = "secret key"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

app.config['MAX_CONTENT_LENGTH'] = 16*1024*1024
basedir = os.path.abspath(os.path.dirname(__file__))

ALLOWED_EXTENSIONS=set(['txt','png','PNG','jpg','xls','xlsx','DAT','dat','zip'])

def rename_duplicate(list):
    new_list=[v + str(list[:i].count(v) + 1) if list.count(v) > 1 else v for i, v in enumerate(list)]
    return new_list

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1] in ALLOWED_EXTENSIONS

def get_FileSize(filePath):
    filePath = unicode(filePath,'utf8')
    fsize = os.path.getsize(filePath)
    fsize = fsize/float(1024*1024)
    return round(fsize,2)

@app.route('/')
def upload_test():
    return render_template('Upload.html')


@app.route('/',methods=['GET','POST'])
def api_upload():
    file = request.files['myfile']
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        z = zipfile.ZipFile(os.path.join(app.config['UPLOAD_FOLDER'], filename), "r")
        z.extractall(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4]))
        z.close()
        for f in os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4])):
            if f[-4:] == '.zip':
                z = zipfile.ZipFile(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4], f), "r")
                z.extractall(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4]))
                z.close()
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.prd')
                data_cfg = open(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.cfg', mode='r')
                data_cfg = pd.DataFrame(data_cfg)
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.cfg')
                x = 11
                y = []
                while data_cfg[0][x][:2] != '\n':
                    y.append(data_cfg[0][x][:2])
                    x = x + 1
                #print(y)
                for x in range(0,len(y)):
                    if y[x] == '05':
                        y[x] = 'VB'
                    elif y[x] == '21':
                        y[x] = 'Delta'
                    elif y[x] == '06':
                        y[x] = 'IR'
                    elif y[x] == '02':
                        y[x] = 'TRR'
                    elif y[x] == '23':
                        y[x] = 'TRR OFFSET'
                    elif y[x] == '01':
                        y[x] = 'KELVIN'
                    elif y[x] == '22':
                        y[x] = 'ABS'
                    elif y[x] == '04':
                        y[x] = 'VF(A)'
                y = rename_duplicate(y)
                for x in range(0,len(y)):
                    y[x] = y[x] + ' NG'
                dbf = DBF(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.dbf')
                frame = pd.DataFrame(iter(dbf))
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.dbf')
                
                frame['X          '] = (-1)*frame['X          ']
                frame['X          '] = frame['X          '] - frame['X          '].min()
                frame['Y          '] = frame['Y          '] - frame['Y          '].min()
                frame['X          '] = frame['X          '] + 1+1
                frame['Y          '] = frame['Y          '] + 1
                data_bin = pd.DataFrame(columns=range(0,int(frame['X          '].max())+1),index = range(0,frame.shape[0]))
                for x in range(0,frame.shape[0]):
                    #print(frame['BIN        '][x])
                    data_bin[frame['X          '][x]][frame['Y          '][x]] = frame['BIN        '][x]
                
                final_place = os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4]) + '.xlsx'
                data_bin.to_excel(final_place,index=False,header=None)
                wb = openpyxl.load_workbook(final_place)
                ws = wb.active
                for col in range(1,ws.max_column+1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4
                    for row in range(1,ws.max_row+1):
                        if ws.cell(row,col).value == None:
                            0
                        elif ws.cell(row,col).value == 81:
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                        elif y[ws.cell(row,col).value-1] == 'VB NG' or y[ws.cell(row,col).value-1] == 'VB1 NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(row,col).value-1] == 'VB2 NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(row,col).value-1] == 'Delta NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(row,col).value-1] == 'IR NG' or y[ws.cell(row,col).value-1] == 'IR1 NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(row,col).value-1] == 'IR2 NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(row,col).value-1] == 'TRR NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(row,col).value-1] == 'TRR OFFSET NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(row,col).value-1] == 'KELVIN NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(row,col).value-1] == 'ABS NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(row,col).value-1] == 'VF(A) NG':
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                ws.cell(row=1, column=1, value=f[:-4])
                if ws.cell(1,3).value != None:
                    ws.cell(2,2).value = 'GOOD'
                    ws.cell(2,1).value = 81
                    ws.cell(2,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+2+1,1).value = x+1
                        if y[ws.cell(x+2+1,1).value-1] == 'VB NG' or y[ws.cell(x+2+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+2+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+2+1,1).value-1] == 'Delta NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+2+1,1).value-1] == 'IR NG' or y[ws.cell(x+2+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+2+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+2+1,1).value-1] == 'TRR NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+2+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+2+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+2+1,1).value-1] == 'ABS NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+2+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+2+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+2+1,2).value = y[x]
                elif ws.cell(2,6).value != None:
                    ws.merge_cells('A1:C1')
                    ws.cell(3,2).value = 'GOOD'
                    ws.cell(3,1).value = 81
                    ws.cell(3,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+3+1,1).value = x+1
                        if y[ws.cell(x+3+1,1).value-1] == 'VB NG' or y[ws.cell(x+3+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+3+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+3+1,1).value-1] == 'Delta NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+3+1,1).value-1] == 'IR NG' or y[ws.cell(x+3+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+3+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+3+1,1).value-1] == 'TRR NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+3+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+3+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+3+1,1).value-1] == 'ABS NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+3+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+3+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+3+1,2).value = y[x]
                elif ws.cell(3,9).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=16, bold=True)
                    ws.merge_cells('A1:F2')
                    ws.cell(4,2).value = 'GOOD'
                    ws.cell(4,1).value = 81
                    ws.cell(4,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+4+1,1).value = x+1
                        if y[ws.cell(x+4+1,1).value-1] == 'VB NG' or y[ws.cell(x+4+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+4+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+4+1,1).value-1] == 'Delta NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+4+1,1).value-1] == 'IR NG' or y[ws.cell(x+4+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+4+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+4+1,1).value-1] == 'TRR NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+4+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+4+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+4+1,1).value-1] == 'ABS NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+4+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+4+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+4+1,2).value = y[x]
                elif ws.cell(4,12).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=24, bold=True)
                    ws.merge_cells('A1:I3')
                    ws.cell(5,2).value = 'GOOD'
                    ws.cell(5,1).value = 81
                    ws.cell(5,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+5+1,1).value = x+1
                        if y[ws.cell(x+5+1,1).value-1] == 'VB NG' or y[ws.cell(x+5+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+5+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+5+1,1).value-1] == 'Delta NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+5+1,1).value-1] == 'IR NG' or y[ws.cell(x+5+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+5+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+5+1,1).value-1] == 'TRR NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+5+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+5+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+5+1,1).value-1] == 'ABS NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+5+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+5+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+5+1,2).value = y[x]
                elif ws.cell(5,15).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=30, bold=True)
                    ws.merge_cells('A1:L4')
                    ws.cell(6,2).value = 'GOOD'
                    ws.cell(6,1).value = 81
                    ws.cell(6,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+6+1,1).value = x+1
                        if y[ws.cell(x+6+1,1).value-1] == 'VB NG' or y[ws.cell(x+6+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+6+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+6+1,1).value-1] == 'Delta NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+6+1,1).value-1] == 'IR NG' or y[ws.cell(x+6+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+6+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+6+1,1).value-1] == 'TRR NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+6+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+6+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+6+1,1).value-1] == 'ABS NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+6+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+6+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+6+1,2).value = y[x]
                elif ws.cell(6,18).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=40, bold=True)
                    ws.merge_cells('A1:O5')
                    ws.cell(7,2).value = 'GOOD'
                    ws.cell(7,1).value = 81
                    ws.cell(7,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+7+1,1).value = x+1
                        if y[ws.cell(x+7+1,1).value-1] == 'VB NG' or y[ws.cell(x+7+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+7+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+7+1,1).value-1] == 'Delta NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+7+1,1).value-1] == 'IR NG' or y[ws.cell(x+7+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+7+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+7+1,1).value-1] == 'TRR NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+7+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+7+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+7+1,1).value-1] == 'ABS NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+7+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+7+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+7+1,2).value = y[x]
                else:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=48, bold=True)
                    ws.merge_cells('A1:R6')
                    ws.cell(8,2).value = 'GOOD'
                    ws.cell(8,1).value = 81
                    ws.cell(8,1).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                    for x in range(0,len(y)):
                        ws.cell(x+8+1,1).value = x+1
                        if y[ws.cell(x+8+1,1).value-1] == 'VB NG' or y[ws.cell(x+8+1,1).value-1] == 'VB1 NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                        elif y[ws.cell(x+8+1,1).value-1] == 'VB2 NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='F75000')
                        elif y[ws.cell(x+8+1,1).value-1] == 'Delta NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='B87070')
                        elif y[ws.cell(x+8+1,1).value-1] == 'IR NG' or y[ws.cell(x+8+1,1).value-1] == 'IR1 NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0080')
                        elif y[ws.cell(x+8+1,1).value-1] == 'IR2 NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF00FF')
                        elif y[ws.cell(x+8+1,1).value-1] == 'TRR NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='FF0000')
                        elif y[ws.cell(x+8+1,1).value-1] == 'TRR OFFSET NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='00BB00')
                        elif y[ws.cell(x+8+1,1).value-1] == 'KELVIN NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='2828FF')
                        elif y[ws.cell(x+8+1,1).value-1] == 'ABS NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='5A5AAD')
                        elif y[ws.cell(x+8+1,1).value-1] == 'VF(A) NG':
                            ws.cell(x+8+1,1).fill = openpyxl.styles.PatternFill('solid', fgColor='796400')
                        ws.cell(x+8+1,2).value = y[x]
                wb.save(final_place)
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f))
            if f[-4:] == '.DAT'or f[-4:] == '.dat':
                file_test = open(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f),'r')
                file_test2 = open(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4] + '.csv'),'w')
                for lines in file_test.readlines():
                    strdata = ",".join(lines.split('\t'))
                    file_test2.write(strdata)
                file_test.close()
                file_test2.close()
                data = pd.read_csv(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4] + '.csv'),header = None)
                for x in range(0,data.shape[0]):
                    if data[0][x] == "[DATA32]":
                        newdata = data[0][x+1]
                        break
                a = 0
                b = []
                for x in range(0,len(newdata)):
                    if newdata[x] == '=' and x != 0:
                        #print(newdata[a:x])
                        #finaldata.insert(newdata[a:x])
                        b.append(newdata[a:x])
                        a = x
                if a == 0:
                    b.append(newdata)
                a = pd.DataFrame(b)
                a = a[0].str.split('|',expand=True)
                a.rename(columns={0:'NO',1:'BIN',2:'unknow',3:'VF1',4:'VF2',5:'SG1',6:'VR1',7:'VR2',8:'dVR1',9:'IR1',10:'dIR1',11:'VR3',12:'dVR2',13:'TRR1',14:'SOP1'},inplace=True)
                t = a['NO'].str.split(' ',expand=True)
                t01 = t[0].str.split(':',expand=True)
                t01.rename(columns={0:'NO',1:'X'},inplace=True)
                t02 = t[1].str.split(':',expand=True)
                t02.rename(columns={0:'NO',1:'Y'},inplace=True)
                newdata = pd.concat([a,t01['X'],t02['Y']], axis=1)
                newdata['XX'] = newdata['X']
                newdata['XX'] = newdata['XX'].str.replace("-","")
                #newdata.to_csv('C:\\Users\\minghsuantu\\Desktop\\check_1.csv',index=0,header=0)
                for x in range(0,newdata.shape[0]):
                    if newdata['XX'][x] == "":
                        newdata = newdata.drop(x)
                newdata = newdata.reset_index(drop=True)
                newdata['X'] = newdata['X'].astype(int)
                newdata['Y'] = newdata['Y'].astype(int)
                newdata['X'] = newdata['X'] - newdata['X'].min()
                newdata['Y'] = newdata['Y'] - newdata['Y'].min()
                newdata['X'] = newdata['X'] + 1
                newdata['Y'] = newdata['Y'] + 1
                newdata['BIN'] = newdata['BIN'].astype(int)
                #newdata.to_csv('C:\\Users\\minghsuantu\\Desktop\\check.csv',index=0,header=0)
                data_bin = pd.DataFrame(columns=range(0,int(newdata['X'].max())+1),index = range(0,newdata.shape[0]))
                for x in range(0,newdata.shape[0]):
                    #print(frame['BIN        '][x])
                    data_bin[newdata['X'][x]][newdata['Y'][x]] = newdata['BIN'][x]
                final_place = os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],str.split(f[:-4])[0]) + '.xlsx'
                data_bin.to_excel(final_place,index=False,header=None)
                wb = openpyxl.load_workbook(final_place)
                ws = wb.active
                for col in range(1,ws.max_column+1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4
                    for row in range(1,ws.max_row+1):
                        if ws.cell(row,col).value == None:
                            0
                        elif ws.cell(row,col).value == 1:
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='80FFFF')
                        else:
                            ws.cell(row,col).fill = openpyxl.styles.PatternFill('solid', fgColor='FFA500')
                ws.cell(row=1, column=1, value=str.split(f[:-4])[0])
                if ws.cell(1,3).value != None:
                    0
                elif ws.cell(2,6).value != None:
                    ws.merge_cells('A1:C1')
                elif ws.cell(3,9).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=16, bold=True)
                    ws.merge_cells('A1:F2')
                elif ws.cell(4,12).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=24, bold=True)
                    ws.merge_cells('A1:I3')
                elif ws.cell(5,15).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=30, bold=True)
                    ws.merge_cells('A1:L4')
                elif ws.cell(6,18).value != None:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=40, bold=True)
                    ws.merge_cells('A1:O5')
                else:
                    ws['A1'].font = openpyxl.styles.Font(name='TIMES NEW Roman', size=48, bold=True)
                    ws.merge_cells('A1:R6')
                        
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f))
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4],f[:-4] + '.csv'))
                wb.save(final_place)
        os.chdir(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4]))
        ll = os.listdir()
        z = zipfile.ZipFile(filename[:-4]+'_plot.zip', 'w', zipfile.ZIP_DEFLATED)
        for w in ll:
            z.write(w)
            os.remove(w)
        z.close()
        shutil.move(filename[:-4]+'_plot.zip',app.config['UPLOAD_FOLDER'])
        os.chdir(app.config['UPLOAD_FOLDER'])
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        os.removedirs(os.path.join(app.config['UPLOAD_FOLDER'], filename[:-4]))
        
        flash('File successfully uploaded')
        return redirect('/')
    

def getListFiles(path):
    ret=[]
    for root,dirs,files in os.walk(path):
        for filespath in files:
            ret.append(os.path.join(filespath))
    return ret

@app.route("/download")
def download_list():
    file_dir=os.path.join(basedir,app.config['UPLOAD_FOLDER'])
    f_list = getListFiles(file_dir)
    len_file=len(f_list)
    return render_template('Download_list.html',f_list=f_list,len_file=len_file)


@app.route("/download/<filename>",methods=['GET'])
def download(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],filename,as_attachment=True)

@app.route("/delete/<filename>",methods=['GET'])
def delet(filename):
    file_dir = os.path.join(basedir,app.config['UPLOAD_FOLDER'])
    delfile = os.path.join(file_dir,filename)
    if os.path.exists(delfile):
        os.remove(delfile)
        return "ok<a href=\"..\download\">File List</a>"
    else:
        return "delete fail"



if __name__ == '__main__':
    app.run(debug=True)